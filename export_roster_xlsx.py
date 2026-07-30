"""把仲介名冊 JSON 匯出成 Excel,供人工查看／查號。

    python export_roster_xlsx.py                    # → data/agency_roster.xlsx
    python export_roster_xlsx.py 我的名冊.xlsx        # 指定輸出路徑

名冊更新(重新下載 data/agency_roster.json)後直接重跑即可,不需改程式。

刻意「不」用 Excel 直接開 JSON／CSV:官方欄位裡許可證(0001、0002-1)與統一編號
都是前導零字串,Excel 會吃掉前導零、還會把 1-73 判讀成日期(Jan-73)。本腳本把
這兩欄寫成文字格式(@),原樣保留。

額外加三個「衍生欄」,都直接沿用 permit_lookup 的同一套判斷,確保人工看到的
與 pipeline 反查時用的完全一致(不另寫一套規則):
  - 狀態      ← _is_active():未廢止且未終止營業 = 有效
  - 電話(正規化) ← _norm_phone():去分隔符、+886 還原前導 0,這才是反查用的鍵
  - 電話備註   ← 標出反查會失準的列(同欄多組號碼/分機、碼數不足)
"""
from __future__ import annotations

import logging
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 私有函式刻意直接引用:衍生欄的意義就是「與 pipeline 反查同一套規則」,
# 若在此重寫一份,名冊判斷與實際反查就會各走各的。
from permit_lookup import (ROSTER_PATH, _is_active, fetch_roster,
                           split_phone_field)

logger = logging.getLogger(__name__)

OUT_PATH = ROSTER_PATH.with_suffix(".xlsx")

# 官方 15 欄的顯示順序:常查的(許可證/名稱/地址/電話)排前面,日期類集中在後。
# 衍生欄插在它所解釋的原始欄旁邊。
COLUMNS: list[tuple[str, int]] = [   # (欄名, 欄寬)
    ("許可證", 10),
    ("機構名稱", 30),
    ("機構地址", 46),
    ("電話", 20),
    ("電話(正規化)", 16),
    ("電話備註", 16),
    ("狀態", 9),
    ("負責人姓名", 12),
    ("公司統一編號", 13),
    ("專業人員人數", 11),
    ("從業人員人數", 11),
    ("許可證起始日", 12),
    ("許可證終止日", 12),
    ("停業起始日", 12),
    ("停業屆滿日", 12),
    ("預訂復業日期", 12),
    ("終止營業日期", 12),
    ("廢止許可日期", 12),
]
TEXT_COLS = {"許可證", "公司統一編號"}       # 需保留前導零 → 文字格式
DATE_COLS = {c for c, _ in COLUMNS if c.endswith(("日", "日期"))}

_MULTI_SEP = re.compile(r"[/、.#*]")   # 同一欄擠進兩支號碼(/、.)或接分機(#*)


def _fmt_date(v: str) -> str:
    """官方日期是 8 碼西元字串(20061020)→ 2006-10-20;非 8 碼數字則原樣回傳。"""
    s = str(v or "").strip()
    if len(s) == 8 and s.isdigit():
        return f"{s[:4]}-{s[4:6]}-{s[6:]}"
    return s


def phone_note(raw: str) -> str:
    """標註這格電話的形態,讓人工看得懂「電話(正規化)」為何是那個樣子。

    多組/分機:permit_lookup.split_phone_field 會拆成多把 key(兩支號碼各建一把、
      分機切掉),所以兩支都查得到——此註記是說明,不是問題。
    碼數不足:該號碼進不了索引(台灣市話/手機正規化後 9~10 碼)。實測這類
      全是已終止/廢止的舊資料,本來就不在反查範圍內。
    """
    raw = str(raw or "").strip()
    if not raw:
        return "無電話"
    keys = split_phone_field(raw)
    if not keys:
        return "碼數不足"
    if _MULTI_SEP.search(raw) or len(keys) > 1:
        return "多組或含分機(已拆分)"
    return ""


def to_row(rec: dict) -> list[str]:
    """一筆名冊記錄 → 對齊 COLUMNS 的一列值(全部為字串,避免 Excel 亂轉型)。"""
    phone = str(rec.get("電話", "")).strip()
    derived = {
        # 一格兩支號碼時會有兩把 key,全部列出——這才是索引裡真正存在的東西
        "電話(正規化)": "、".join(split_phone_field(phone)),
        "電話備註": phone_note(phone),
        "狀態": "有效" if _is_active(rec) else "已終止/廢止",
    }
    row = []
    for name, _ in COLUMNS:
        if name in derived:
            row.append(derived[name])
        elif name in DATE_COLS:
            row.append(_fmt_date(rec.get(name, "")))
        else:
            row.append(str(rec.get(name, "")).strip())
    return row


def _add_readme(wb: Workbook, records: list[dict], src: Path) -> None:
    """加一張「說明」表:資料來源、匯出時間與幾個一眼可見的統計。"""
    ws = wb.create_sheet("說明")
    notes = [str(phone_note(r.get("電話", ""))) for r in records]
    mtime = datetime.fromtimestamp(src.stat().st_mtime).strftime("%Y-%m-%d")
    lines = [
        ("仲介名冊(供人工查看)", ""),
        ("", ""),
        ("資料來源", "勞動部「跨國人力仲介公司許可名冊」data.gov.tw dataset 6682"),
        ("來源檔", str(src)),
        ("來源檔更新日", mtime),
        ("匯出時間", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("總筆數", f"{len(records)}"),
        ("目前有效", f"{sum(1 for r in records if _is_active(r))}"),
        ("已終止/廢止", f"{sum(1 for r in records if not _is_active(r))}"),
        ("", ""),
        ("電話「多組或含分機(已拆分)」", f"{notes.count('多組或含分機(已拆分)')}"
                                       "(兩支號碼各建一把索引,分機切掉,都查得到)"),
        ("電話「碼數不足」", f"{notes.count('碼數不足')}(進不了索引;實測全是已終止/廢止的舊資料)"),
        ("電話「無電話」", f"{notes.count('無電話')}"),
        ("", ""),
        ("欄位說明", ""),
        ("電話(正規化)", "去非數字、+886 還原前導 0、切掉分機;程式反查比對的就是這一欄。"
                        "一格塞兩支號碼時兩把都列出(以「、」分隔),索引裡確實各有一把"),
        ("電話備註", "空白=單一號碼;其餘見上方統計的說明"),
        ("狀態", "未廢止且未終止營業 = 有效;程式只反查有效的機構"),
        ("許可證/統一編號", "以文字格式保存,前導零與 1-73 這類分支號不會被 Excel 轉掉"),
        ("", ""),
        ("更新方式", "重新下載來源 JSON 覆蓋後,執行 python export_roster_xlsx.py 重匯出"),
    ]
    for k, v in lines:
        ws.append([k, v])
    for row in ws.iter_rows(min_col=1, max_col=1):
        for c in row:
            if c.value:
                c.font = Font(bold=True)
    ws["A1"].font = Font(bold=True, size=14)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 78
    for row in ws.iter_rows(min_col=2, max_col=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")


def export(out_path: Path = OUT_PATH, src: Path = ROSTER_PATH) -> Path:
    """讀名冊 JSON → 寫出 xlsx(名冊 + 說明兩張表);回傳實際寫出的路徑。"""
    records = fetch_roster(src)
    wb = Workbook()
    ws = wb.active
    ws.title = "名冊"

    ws.append([name for name, _ in COLUMNS])
    for rec in records:
        ws.append(to_row(rec))

    # 標題列樣式 + 凍結 + 篩選(人工最常用的三件事)
    head_fill = PatternFill("solid", fgColor="DDEBF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 22

    for i, (name, width) in enumerate(COLUMNS, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = width
        if name in TEXT_COLS:
            for cell in ws[letter][1:]:      # 跳過標題列
                cell.number_format = "@"

    _add_readme(wb, records, src)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else OUT_PATH
    path = export(out)
    logger.info(f"✓ 已匯出:{path}({path.stat().st_size / 1024:.0f} KB)")
